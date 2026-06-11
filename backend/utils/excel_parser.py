import csv
import io
import os
from typing import Any, Dict, List, Union
import openpyxl

try:
    from openpyxl.worksheet.filters import CustomFilterValueDescriptor
    if not getattr(CustomFilterValueDescriptor, "_patched", False):
        from openpyxl.descriptors.base import Convertible
        _orig_set = CustomFilterValueDescriptor.__set__
        def _patched_set(self, instance, value):
            if isinstance(value, str):
                self.expected_type = str
                Convertible.__set__(self, instance, value)
            else:
                _orig_set(self, instance, value)
        CustomFilterValueDescriptor.__set__ = _patched_set
        CustomFilterValueDescriptor._patched = True
except Exception as e:
    print(f"[Monkeypatch Warning] Failed to patch openpyxl filters descriptor: {e}")

from openpyxl.worksheet.worksheet import Worksheet

class RawRow:
    def __init__(self, row_idx: int, cells: List[Any]):
        self.row_idx = row_idx
        self.cells = cells

    def to_dict(self) -> Dict[str, Any]:
        return {
            "row_idx": self.row_idx,
            "cells": self.cells
        }


def _build_merged_lookup(sheet: Worksheet) -> Dict[str, Any]:
    """
    Pre-builds a coordinate → top-left-cell-value mapping for every cell that
    sits inside a merged range.  This lets us do O(1) lookups per cell instead
    of the previous O(cells × merged_ranges) inner loop.
    """
    lookup: Dict[str, Any] = {}
    try:
        for merged_range in sheet.merged_cells.ranges:
            top_left_val = sheet.cell(
                row=merged_range.min_row, column=merged_range.min_col
            ).value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    coord = sheet.cell(row=row, column=col).coordinate
                    lookup[coord] = top_left_val
    except Exception:
        # read_only mode doesn't expose merged_cells — that's fine.
        pass
    return lookup


def parse_excel_sheet(sheet: Worksheet) -> List[RawRow]:
    """
    Parses a single openpyxl worksheet, resolving merged cells.
    Returns a list of RawRow objects with 1-based row indices.
    """
    merged_lookup = _build_merged_lookup(sheet)
    raw_rows = []

    for r in range(1, sheet.max_row + 1):
        cells_values = []
        has_content = False

        for c in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=r, column=c)
            coord = cell.coordinate
            # Use pre-built lookup for merged cells; fall back to own value
            val = merged_lookup.get(coord, cell.value)

            if val is not None:
                if isinstance(val, str):
                    val_str = val.strip()
                    if val_str == "":
                        val = None
                    else:
                        val = val_str
                        has_content = True
                else:
                    has_content = True

            cells_values.append(val)

        raw_rows.append(RawRow(row_idx=r, cells=cells_values))

    return raw_rows

def parse_excel_file(file_path: str) -> Dict[str, List[RawRow]]:
    """
    Loads an Excel workbook and parses each sheet.
    Returns a mapping of sheet_name -> list of RawRows.
    """
    # read_only=False is required to access merged_cells; data_only=True skips formulas.
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
    results = {}
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        results[sheet_name] = parse_excel_sheet(sheet)
    wb.close()
    return results

def parse_csv_file(file_path: str, encoding: str = 'utf-8-sig') -> List[RawRow]:
    """
    Parses a CSV file. Returns a list of RawRows.
    """
    raw_rows = []
    with open(file_path, 'r', encoding=encoding, errors='ignore') as f:
        reader = csv.reader(f)
        for idx, row in enumerate(reader, start=1):
            cells = [cell.strip() if cell.strip() != "" else None for cell in row]
            raw_rows.append(RawRow(row_idx=idx, cells=cells))
    return raw_rows

def parse_file(file_path: str) -> Dict[str, List[RawRow]]:
    """
    Unified parser entry point. Auto-detects Excel vs CSV and parses.
    Returns a dict mapping sheet name (or 'default' for CSV) to List[RawRow].
    """
    ext = os.path.splitext(file_path)[1].lower()
    if ext in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
        return parse_excel_file(file_path)
    elif ext == '.csv':
        return {"default": parse_csv_file(file_path)}
    elif ext == '.xls':
        # openpyxl does not support legacy .xls format.
        # Inform user to save as .xlsx or convert first.
        raise ValueError("Legacy Excel .xls format not supported. Please save as .xlsx or .csv.")
    else:
        raise ValueError(f"Unsupported file format: {ext}")
