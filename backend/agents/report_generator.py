import os
from typing import List, Dict, Any, Optional
from decimal import Decimal
import datetime as dt
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from backend.models.schema import (
    DuplicatePaymentFinding,
    PartyAgingSummary,
    AnomalyFinding,
    ReconciliationPartySummary,
    AuditMemoModel
)
from backend.utils.llm import LLMClient

# Color Palette Definitions (Pastel / Executive Sleek)
NAVY_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
WHITE_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
REGULAR_FONT = Font(name="Calibri", size=11)
TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True)

# Risk Badge Fills
CRITICAL_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # soft red-orange
CRITICAL_FONT = Font(name="Calibri", size=11, bold=True, color="C00000")

HIGH_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
HIGH_FONT = Font(name="Calibri", size=11, bold=True, color="C00000")

MEDIUM_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # soft yellow
MEDIUM_FONT = Font(name="Calibri", size=11, bold=True, color="7F6000")

LOW_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # soft green
LOW_FONT = Font(name="Calibri", size=11, bold=True, color="375623")

# Aging Color-Coding Fills
AGE_GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
AGE_YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
AGE_ORANGE_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
AGE_RED_FILL = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

# Border Definitions
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
DOUBLE_BOTTOM_BORDER = Border(
    top=Side(style='thin', color='1F4E79'),
    bottom=Side(style='double', color='1F4E79')
)

def auto_fit_columns(ws):
    """Auto-fits worksheet column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Skip merged cells calculation where width could be misleading
            if cell.coordinate in ws.merged_cells:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        # Apply width with safety margin
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

def generate_excel_report(
    file_path: str,
    duplicates: List[DuplicatePaymentFinding],
    aging: List[PartyAgingSummary],
    anomalies: List[AnomalyFinding],
    reconciliation: List[ReconciliationPartySummary]
):
    """
    Creates a highly styled, executive-grade multi-sheet audit Excel workbook.
    """
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # ----------------------------------------------------
    # Sheet 1: Executive Summary
    # ----------------------------------------------------
    ws_sum = wb.create_sheet(title="Executive Summary")
    ws_sum.views.sheetView[0].showGridLines = True
    
    ws_sum.cell(row=1, column=1, value="GL Ledger Forensic Audit Report").font = TITLE_FONT
    ws_sum.cell(row=2, column=1, value=f"Generated on: {dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = SUBTITLE_FONT
    
    # KPI Blocks Table
    ws_sum.cell(row=4, column=1, value="Key Performance Indicators (KPIs)").font = BOLD_FONT
    
    kpi_headers = ["Metric Description", "Value"]
    for col_idx, h in enumerate(kpi_headers, start=1):
        cell = ws_sum.cell(row=5, column=col_idx, value=h)
        cell.font = WHITE_FONT
        cell.fill = NAVY_HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        
    total_outstanding = sum(a.outstanding_balance for a in aging)
    dup_risk_val = sum(Decimal(str(d.transaction_A.get("amount", 0.0))) for d in duplicates if d.pass_number != 5)
    critical_anom_count = sum(1 for a in anomalies if a.severity in ("CRITICAL", "HIGH"))
    over_year_outstanding = sum(a.aging_buckets.get(">365", Decimal("0.00")) for a in aging)
    
    kpi_rows = [
        ("Total Audited Parties", len(aging)),
        ("Gross Outstanding Position", float(total_outstanding)),
        ("Duplicate Payments Risk (Pass 1-4)", float(dup_risk_val)),
        ("Critical/High Severity Anomalies", critical_anom_count),
        ("Outstanding Balance Over 1 Year", float(over_year_outstanding))
    ]
    
    for row_idx, (desc, val) in enumerate(kpi_rows, start=6):
        c1 = ws_sum.cell(row=row_idx, column=1, value=desc)
        c2 = ws_sum.cell(row=row_idx, column=2, value=val)
        c1.font = REGULAR_FONT
        c2.font = BOLD_FONT
        c1.border = THIN_BORDER
        c2.border = THIN_BORDER
        
        if isinstance(val, float):
            c2.number_format = "$#,##0.00;($#,##0.00);\"-\""
            c2.alignment = Alignment(horizontal="right")
        else:
            c2.alignment = Alignment(horizontal="center")
            
    # Risk Heatmap summary (Count by anomaly severity)
    ws_sum.cell(row=13, column=1, value="Anomaly Summary by Severity").font = BOLD_FONT
    
    heat_headers = ["Severity Level", "Anomaly Count"]
    for col_idx, h in enumerate(heat_headers, start=1):
        cell = ws_sum.cell(row=14, column=col_idx, value=h)
        cell.font = WHITE_FONT
        cell.fill = NAVY_HEADER_FILL
        
    severity_counts = {
        "CRITICAL": sum(1 for a in anomalies if a.severity == "CRITICAL"),
        "HIGH": sum(1 for a in anomalies if a.severity == "HIGH"),
        "MEDIUM": sum(1 for a in anomalies if a.severity == "MEDIUM"),
        "LOW": sum(1 for a in anomalies if a.severity == "LOW")
    }
    
    for r_idx, (sev, count) in enumerate(severity_counts.items(), start=15):
        c1 = ws_sum.cell(row=r_idx, column=1, value=sev)
        c2 = ws_sum.cell(row=r_idx, column=2, value=count)
        c1.font = BOLD_FONT
        c2.font = BOLD_FONT
        c1.border = THIN_BORDER
        c2.border = THIN_BORDER
        c2.alignment = Alignment(horizontal="center")
        
        # Color coding severity cells
        if sev == "CRITICAL":
            c1.fill = CRITICAL_FILL
            c1.font = CRITICAL_FONT
        elif sev == "HIGH":
            c1.fill = HIGH_FILL
            c1.font = HIGH_FONT
        elif sev == "MEDIUM":
            c1.fill = MEDIUM_FILL
            c1.font = MEDIUM_FONT
        elif sev == "LOW":
            c1.fill = LOW_FILL
            c1.font = LOW_FONT

    auto_fit_columns(ws_sum)

    # ----------------------------------------------------
    # Sheet 2: Aging Schedule
    # ----------------------------------------------------
    ws_age = wb.create_sheet(title="Aging Schedule")
    ws_age.views.sheetView[0].showGridLines = True
    
    ws_age.cell(row=1, column=1, value="Outstanding Aging Schedule").font = TITLE_FONT
    
    age_headers = ["Party Name", "Opening Balance", "Total Debits", "Total Credits", "Net Outstanding", "0-30 days", "31-60 days", "61-90 days", "91-180 days", "181-365 days", ">365 days", "Flags"]
    for col_idx, h in enumerate(age_headers, start=1):
        cell = ws_age.cell(row=3, column=col_idx, value=h)
        cell.font = WHITE_FONT
        cell.fill = NAVY_HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        
    for r_idx, ag in enumerate(aging, start=4):
        ws_age.cell(row=r_idx, column=1, value=ag.party).font = REGULAR_FONT
        
        # Values
        vals = [
            ag.opening_balance, ag.total_debits, ag.total_credits, ag.outstanding_balance,
            ag.aging_buckets.get("0-30", Decimal("0.00")),
            ag.aging_buckets.get("31-60", Decimal("0.00")),
            ag.aging_buckets.get("61-90", Decimal("0.00")),
            ag.aging_buckets.get("91-180", Decimal("0.00")),
            ag.aging_buckets.get("181-365", Decimal("0.00")),
            ag.aging_buckets.get(">365", Decimal("0.00"))
        ]
        
        for c_idx, val in enumerate(vals, start=2):
            cell = ws_age.cell(row=r_idx, column=c_idx, value=float(val))
            cell.font = REGULAR_FONT
            cell.border = THIN_BORDER
            cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right")
            
            # Color code aging columns
            # Column 6 (0-30) = green, Column 7/8 (31-90) = yellow, Column 9 (91-180) = orange, Column 10/11 (>180) = red
            if val > 0:
                if c_idx == 6:
                    cell.fill = AGE_GREEN_FILL
                elif c_idx in (7, 8):
                    cell.fill = AGE_YELLOW_FILL
                elif c_idx == 9:
                    cell.fill = AGE_ORANGE_FILL
                elif c_idx in (10, 11):
                    cell.fill = AGE_RED_FILL
                    
        # Flags
        flags = []
        if ag.flag_unsettled_opening:
            flags.append("Unsettled Opening")
        if ag.flag_zero_payments:
            flags.append("Zero Payments")
            
        flag_cell = ws_age.cell(row=r_idx, column=12, value=", ".join(flags))
        flag_cell.font = BOLD_FONT
        flag_cell.border = THIN_BORDER
        if flags:
            flag_cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            flag_cell.font = Font(name="Calibri", size=11, color="C00000", bold=True)
            
    auto_fit_columns(ws_age)

    # ----------------------------------------------------
    # Sheet 3: Duplicate Payments
    # ----------------------------------------------------
    ws_dup = wb.create_sheet(title="Duplicate Payments")
    ws_dup.views.sheetView[0].showGridLines = True
    
    ws_dup.cell(row=1, column=1, value="Duplicate Payments Analysis").font = TITLE_FONT
    
    dup_headers = [
        "Party", "Pass", "Detection Pass Name", 
        "Txn A Date", "Txn A Voucher", "Txn A Amount", "Txn A Contra",
        "Txn B Date", "Txn B Voucher", "Txn B Amount", "Txn B Contra",
        "Delta Days", "Delta Amount", "Confidence", "Recommendation"
    ]
    for col_idx, h in enumerate(dup_headers, start=1):
        cell = ws_dup.cell(row=3, column=col_idx, value=h)
        cell.font = WHITE_FONT
        cell.fill = NAVY_HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        
    for r_idx, d in enumerate(duplicates, start=4):
        ws_dup.cell(row=r_idx, column=1, value=d.party).font = REGULAR_FONT
        ws_dup.cell(row=r_idx, column=2, value=d.pass_number).font = REGULAR_FONT
        ws_dup.cell(row=r_idx, column=3, value=d.pass_name).font = REGULAR_FONT
        
        # Txn A
        ws_dup.cell(row=r_idx, column=4, value=d.transaction_A.get("date")).font = REGULAR_FONT
        ws_dup.cell(row=r_idx, column=5, value=d.transaction_A.get("voucher_no")).font = REGULAR_FONT
        
        c6 = ws_dup.cell(row=r_idx, column=6, value=d.transaction_A.get("amount"))
        c6.number_format = "#,##0.00"
        c6.alignment = Alignment(horizontal="right")
        
        ws_dup.cell(row=r_idx, column=7, value=d.transaction_A.get("contra_ledger")).font = REGULAR_FONT
        
        # Txn B
        ws_dup.cell(row=r_idx, column=8, value=d.transaction_B.get("date", "")).font = REGULAR_FONT
        ws_dup.cell(row=r_idx, column=9, value=d.transaction_B.get("voucher_no", "")).font = REGULAR_FONT
        
        c10 = ws_dup.cell(row=r_idx, column=10, value=d.transaction_B.get("amount", 0.0))
        c10.number_format = "#,##0.00"
        c10.alignment = Alignment(horizontal="right")
        
        ws_dup.cell(row=r_idx, column=11, value=d.transaction_B.get("contra_ledger", "")).font = REGULAR_FONT
        
        # Deltas
        c12 = ws_dup.cell(row=r_idx, column=12, value=d.delta_days)
        c12.alignment = Alignment(horizontal="center")
        
        c13 = ws_dup.cell(row=r_idx, column=13, value=float(d.delta_amount))
        c13.number_format = "#,##0.00"
        c13.alignment = Alignment(horizontal="right")
        
        # Confidence Badge
        conf_cell = ws_dup.cell(row=r_idx, column=14, value=d.confidence)
        conf_cell.alignment = Alignment(horizontal="center")
        if d.confidence == "HIGH":
            conf_cell.fill = CRITICAL_FILL
            conf_cell.font = CRITICAL_FONT
        elif d.confidence == "MEDIUM":
            conf_cell.fill = MEDIUM_FILL
            conf_cell.font = MEDIUM_FONT
        else:
            conf_cell.fill = LOW_FILL
            conf_cell.font = LOW_FONT
            
        ws_dup.cell(row=r_idx, column=15, value=d.recommendation).font = REGULAR_FONT
        
        # Add border to all cells in row
        for c in range(1, 16):
            ws_dup.cell(row=r_idx, column=c).border = THIN_BORDER
            
    auto_fit_columns(ws_dup)

    # ----------------------------------------------------
    # Sheet 4: Anomaly Register
    # ----------------------------------------------------
    ws_anom = wb.create_sheet(title="Anomaly Register")
    ws_anom.views.sheetView[0].showGridLines = True
    
    ws_anom.cell(row=1, column=1, value="Forensic Anomaly Register").font = TITLE_FONT
    
    anom_headers = ["Party Name", "Anomaly Type", "Severity", "Detailed Description", "Source Evidence Reference", "Auditor Action Recommendation"]
    for col_idx, h in enumerate(anom_headers, start=1):
        cell = ws_anom.cell(row=3, column=col_idx, value=h)
        cell.font = WHITE_FONT
        cell.fill = NAVY_HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        
    for r_idx, an in enumerate(anomalies, start=4):
        ws_anom.cell(row=r_idx, column=1, value=an.party).font = REGULAR_FONT
        ws_anom.cell(row=r_idx, column=2, value=an.anomaly_type).font = BOLD_FONT
        
        sev_cell = ws_anom.cell(row=r_idx, column=3, value=an.severity)
        sev_cell.alignment = Alignment(horizontal="center")
        if an.severity == "CRITICAL":
            sev_cell.fill = CRITICAL_FILL
            sev_cell.font = CRITICAL_FONT
        elif an.severity == "HIGH":
            sev_cell.fill = HIGH_FILL
            sev_cell.font = HIGH_FONT
        elif an.severity == "MEDIUM":
            sev_cell.fill = MEDIUM_FILL
            sev_cell.font = MEDIUM_FONT
        else:
            sev_cell.fill = LOW_FILL
            sev_cell.font = LOW_FONT
            
        ws_anom.cell(row=r_idx, column=4, value=an.description).font = REGULAR_FONT
        ws_anom.cell(row=r_idx, column=5, value=str(an.evidence)).font = REGULAR_FONT
        ws_anom.cell(row=r_idx, column=6, value=an.recommendation).font = REGULAR_FONT
        
        for c in range(1, 7):
            ws_anom.cell(row=r_idx, column=c).border = THIN_BORDER
            
    auto_fit_columns(ws_anom)

    # ----------------------------------------------------
    # Sheet 5: Reconciliation Statement
    # ----------------------------------------------------
    ws_rec = wb.create_sheet(title="Reconciliation Statement")
    ws_rec.views.sheetView[0].showGridLines = True
    
    ws_rec.cell(row=1, column=1, value="Party ledger Reconciliation Statement").font = TITLE_FONT
    
    rec_headers = ["Party Name", "Stated Opening", "Stated Closing", "Recalculated Closing", "Variance", "Total Debits", "Total Credits", "Status"]
    for col_idx, h in enumerate(rec_headers, start=1):
        cell = ws_rec.cell(row=3, column=col_idx, value=h)
        cell.font = WHITE_FONT
        cell.fill = NAVY_HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        
    for r_idx, rc in enumerate(reconciliation, start=4):
        ws_rec.cell(row=r_idx, column=1, value=rc.party).font = REGULAR_FONT
        
        vals = [rc.stated_opening, rc.stated_closing, rc.recalculated_closing, rc.variance, rc.total_debits, rc.total_credits]
        for c_idx, val in enumerate(vals, start=2):
            cell = ws_rec.cell(row=r_idx, column=c_idx, value=float(val))
            cell.font = REGULAR_FONT
            cell.border = THIN_BORDER
            cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right")
            
        status_cell = ws_rec.cell(row=r_idx, column=8, value="Discrepancy" if rc.has_discrepancy else "Matched")
        status_cell.alignment = Alignment(horizontal="center")
        status_cell.font = BOLD_FONT
        status_cell.border = THIN_BORDER
        
        if rc.has_discrepancy:
            status_cell.fill = CRITICAL_FILL
            status_cell.font = CRITICAL_FONT
        else:
            status_cell.fill = LOW_FILL
            status_cell.font = LOW_FONT
            
    auto_fit_columns(ws_rec)
    
    # Save Workbook
    wb.save(file_path)

def _build_template_memo(
    duplicates: List[DuplicatePaymentFinding],
    aging: List[PartyAgingSummary],
    anomalies: List[AnomalyFinding],
    reconciliation: List[ReconciliationPartySummary],
) -> AuditMemoModel:
    """
    Builds a complete AuditMemoModel from structured findings data without an LLM call.
    Used as a fallback when the model is unavailable or times out.
    """
    from decimal import Decimal
    from datetime import date

    total_outstanding = sum(a.outstanding_balance for a in aging)
    critical_anomalies = [a for a in anomalies if a.severity == "CRITICAL"]
    high_anomalies = [a for a in anomalies if a.severity == "HIGH"]
    rec_discrepancies = [r for r in reconciliation if r.has_discrepancy]
    overdue_parties = [a for a in aging if a.bucket_91_180 + a.bucket_181_365 + a.bucket_over_365 > 0]

    scope = (
        f"Forensic ledger audit covering {len(aging)} parties. "
        f"Net outstanding balance: {total_outstanding:,.2f}. "
        f"Audit date: {date.today().isoformat()}."
    )

    methodology = (
        "Column structure was auto-detected using rule-based heuristics with LLM fallback. "
        "Transactions were normalised and validated for sign, date, and amount integrity. "
        "Duplicate payments were scanned across 5 detection passes (exact match, voucher reference, "
        "same-day same-amount, fuzzy ±1%, and unsupported round numbers). "
        "FIFO cash-matching was applied to compute 6-bucket receivable/payable aging. "
        "An independent running-balance reconciliation was performed step-by-step from the opening balance. "
        "Ten forensic anomaly patterns were evaluated (ghost payments, holiday transactions, split-payment "
        "limit bypasses, segregation-of-duties concentrations, round-trip transactions, same-day reversals, etc.)."
    )

    key_findings = []
    if duplicates:
        key_findings.append(
            f"{len(duplicates)} potential duplicate payment(s) identified across multiple detection passes."
        )
    if rec_discrepancies:
        total_var = sum(abs(r.variance) for r in rec_discrepancies)
        key_findings.append(
            f"{len(rec_discrepancies)} reconciliation discrepancy/discrepancies found; "
            f"cumulative variance {total_var:,.2f}."
        )
    if critical_anomalies:
        types = ", ".join(sorted({a.anomaly_type for a in critical_anomalies}))
        key_findings.append(f"{len(critical_anomalies)} CRITICAL anomaly/anomalies: {types}.")
    if high_anomalies:
        key_findings.append(f"{len(high_anomalies)} HIGH-severity anomaly/anomalies detected.")
    if overdue_parties:
        key_findings.append(
            f"{len(overdue_parties)} party/parties with balances overdue beyond 90 days."
        )
    if not key_findings:
        key_findings.append("No material findings detected in this ledger.")

    risk_ratings: dict = {}
    if critical_anomalies:
        risk_ratings["Overall Risk"] = "CRITICAL"
    elif high_anomalies or rec_discrepancies:
        risk_ratings["Overall Risk"] = "HIGH"
    elif duplicates or overdue_parties:
        risk_ratings["Overall Risk"] = "MEDIUM"
    else:
        risk_ratings["Overall Risk"] = "LOW"
    if duplicates:
        risk_ratings["Duplicate Payments"] = "HIGH" if len(duplicates) > 5 else "MEDIUM"
    if rec_discrepancies:
        risk_ratings["Reconciliation"] = "CRITICAL" if len(rec_discrepancies) > 2 else "HIGH"
    if critical_anomalies:
        risk_ratings["Fraud Risk"] = "CRITICAL"

    recommendations = [
        "Review all flagged duplicate transactions with the accounts payable team and obtain supporting invoices.",
        "Investigate reconciliation variances by tracing each discrepant entry to its source voucher.",
    ]
    if critical_anomalies:
        recommendations.append(
            "Escalate CRITICAL anomalies (e.g., holiday transactions, ghost payments) to management and internal audit immediately."
        )
    if overdue_parties:
        recommendations.append(
            "Initiate recovery proceedings for balances overdue beyond 180 days; review credit limits."
        )
    recommendations.append(
        "Implement dual-authorisation controls for payments above the approval threshold to prevent limit-bypass splits."
    )
    recommendations.append(
        "Schedule a quarterly ledger reconciliation to catch discrepancies before year-end close."
    )

    # Build plain-text markdown body
    lines = ["# Forensic Audit Observation Memo", ""]
    lines += ["## 1. Executive Summary & Audit Scope", scope, ""]
    lines += ["## 2. Audit Methodology", methodology, ""]
    lines += ["## 3. Key Findings"]
    for f in key_findings:
        lines.append(f"- {f}")
    lines.append("")
    lines += ["## 4. Risk Ratings"]
    for k, v in risk_ratings.items():
        lines.append(f"- **{k}**: {v}")
    lines.append("")
    lines += ["## 5. Management Recommendations"]
    for i, r in enumerate(recommendations, 1):
        lines.append(f"{i}. {r}")
    lines.append("")
    lines.append("*Note: This memo was auto-generated from structured audit findings. LLM-assisted narrative was unavailable.*")

    return AuditMemoModel(
        scope=scope,
        methodology=methodology,
        key_findings=key_findings,
        risk_ratings=risk_ratings,
        recommendations=recommendations,
        raw_markdown="\n".join(lines),
    )


# Memo generation timeout — writing a full CA audit memo requires many output tokens;
# give it 90s before falling back to the template-based memo builder.
_MEMO_TIMEOUT_SECS = 90


def generate_audit_memo_markdown(
    duplicates: List[DuplicatePaymentFinding],
    aging: List[PartyAgingSummary],
    anomalies: List[AnomalyFinding],
    reconciliation: List[ReconciliationPartySummary],
    llm_client: LLMClient
) -> AuditMemoModel:
    """
    Attempts to generate a CA-style audit memo via the configured LLM (90s timeout).
    Falls back to a deterministic template-based memo if the model is unavailable or too slow.
    """
    # Create text summary of findings to fit inside prompt context
    dup_summary = []
    for d in duplicates[:15]:
        dup_summary.append(f"- Party: {d.party}, Pass {d.pass_number} ({d.pass_name}): Txn A Row {d.transaction_A.get('row_idx')} (Amt {d.transaction_A.get('amount')}) and Txn B Row {d.transaction_B.get('row_idx', '')} (Amt {d.transaction_B.get('amount', '')}) delta_days={d.delta_days}. Conf: {d.confidence}.")
        
    anom_summary = []
    for a in anomalies[:25]:
        anom_summary.append(f"- Party: {a.party}, Type: {a.anomaly_type}, Severity: {a.severity}. Description: {a.description}. Evidence: {a.evidence}.")
        
    rec_summary = []
    for r in reconciliation:
        if r.has_discrepancy or r.variance > 0:
            rec_summary.append(f"- Party: {r.party}, Stated Opening={r.stated_opening}, Stated Closing={r.stated_closing}, Recalculated Closing={r.recalculated_closing}, Variance={r.variance}.")

    total_outstanding = sum(a.outstanding_balance for a in aging)
    
    findings_context = (
        f"OVERALL METADATA:\n"
        f"- Total Parties Audited: {len(aging)}\n"
        f"- Net Outstanding Balance: Rs. {total_outstanding:.2f}\n"
        f"- Total Duplicates Mapped: {len(duplicates)}\n"
        f"- Total Anomalies Detected: {len(anomalies)}\n\n"
        f"DUPLICATE PAYMENT FINDINGS (Preview):\n"
        f"{chr(10).join(dup_summary)}\n\n"
        f"ANOMALY REGISTER (Preview):\n"
        f"{chr(10).join(anom_summary)}\n\n"
        f"RECONCILIATION DISCREPANCIES:\n"
        f"{chr(10).join(rec_summary)}\n"
    )

    system_prompt = (
        "You are an expert Senior Forensic Auditor and Chartered Accountant.\n"
        "You draft highly professional, formal, mathematically accurate, and audit-grade observation memos based on ledger audit findings."
    )
    
    user_prompt = (
        f"Please write a comprehensive Chartered Accountant style Forensic Audit Observation Memo based on the following findings:\n\n"
        f"{findings_context}\n\n"
        f"The memo must be written in formal business English and structured under the following sections:\n"
        f"1. Executive Summary & Audit Scope\n"
        f"2. Audit Methodology Applied (describe structure detection, FIFO bill matching, reconciliation check, and anomaly scanning)\n"
        f"3. Detailed Key Findings (grouped by category: Reconciliation, Duplicates, and Anomalies with clear risk ratings)\n"
        f"4. Risk Assessments & Business Impact (discuss leakage risks, internal control weaknesses)\n"
        f"5. Management Recommendations (concrete operational steps for correcting entries, strengthening controls, and reclaiming funds)\n\n"
        f"Return the parsed structure model fields plus the raw markdown contents in a JSON schema structure matching the AuditMemoModel."
    )

    if llm_client and llm_client.is_configured():
        try:
            return llm_client.call_structured(
                system_prompt=system_prompt,
                user_prompt=user_prompt,
                response_model=AuditMemoModel,
                timeout_secs=_MEMO_TIMEOUT_SECS,
            )
        except Exception as e:
            print(f"[ReportGenerator] LLM memo generation failed ({e}). Using template fallback.")

    # Template fallback — always produces a complete, well-structured memo from the data.
    return _build_template_memo(duplicates, aging, anomalies, reconciliation)

